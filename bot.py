import os
import logging
import threading
import asyncio
import io
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, ConversationHandler, filters, ContextTypes
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

TOKEN = os.environ["BOT_TOKEN"]
ENCARGADA_ID = 6542830130

ESPERANDO_RESPUESTA, ESPERANDO_FECHA, ESPERANDO_NOMBRE, \
ESPERANDO_CEDULA, ESPERANDO_COMPROBANTE, ESPERANDO_MOTIVO = range(6)

class PingHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"Bot activo OK")
    def log_message(self, *args):
        pass

def iniciar_servidor():
    port = int(os.environ.get("PORT", 10000))
    server = HTTPServer(("0.0.0.0", port), PingHandler)
    server.serve_forever()

def generar_excel_pago(datos: dict, usuario: str) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Pago"
    header_fill = PatternFill("solid", fgColor="1F7A4D")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    for col, titulo in enumerate(["Campo", "Detalle"], 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    filas = [
        ("📋 Tipo de reporte",   "PAGO CONFIRMADO ✅"),
        ("🏢 Empresa / Nombre",  datos.get("nombre", "N/A")),
        ("🪪 Cédula / RUC",      datos.get("cedula", "N/A")),
        ("📅 Fecha de pago",     datos.get("fecha", "N/A")),
        ("👤 Usuario Telegram",  f"@{usuario}"),
        ("🕐 Fecha de registro", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("✅ Estado",            "VA A PAGAR"),
    ]
    fill_par = PatternFill("solid", fgColor="E8F5E9")
    for i, (campo, valor) in enumerate(filas, 2):
        ws.cell(row=i, column=1, value=campo).font = Font(bold=True)
        ws.cell(row=i, column=2, value=valor)
        if i % 2 == 0:
            for col in range(1, 3):
                ws.cell(row=i, column=col).fill = fill_par
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def generar_excel_no_pago(datos: dict, usuario: str) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte No Pago"
    header_fill = PatternFill("solid", fgColor="C0392B")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    for col, titulo in enumerate(["Campo", "Detalle"], 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    filas = [
        ("📋 Tipo de reporte",   "NO VA A PAGAR ❌"),
        ("🏢 Empresa / Nombre",  datos.get("nombre", "N/A")),
        ("🪪 Cédula / RUC",      datos.get("cedula", "N/A")),
        ("📝 Motivo",            datos.get("motivo", "N/A")),
        ("👤 Usuario Telegram",  f"@{usuario}"),
        ("🕐 Fecha de registro", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("❌ Estado",            "NO VA A PAGAR"),
    ]
    fill_par = PatternFill("solid", fgColor="FDECEA")
    for i, (campo, valor) in enumerate(filas, 2):
        ws.cell(row=i, column=1, value=campo).font = Font(bold=True)
        ws.cell(row=i, column=2, value=valor)
        if i % 2 == 0:
            for col in range(1, 3):
                ws.cell(row=i, column=col).fill = fill_par
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 *¡Bienvenido al Bot de Cobranzas Caro!*\n\n"
        "Estamos aquí para ayudarle a gestionar su pago de forma rápida y sencilla. 😊",
        parse_mode="Markdown"
    )
    teclado = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ SI", callback_data="si"),
        InlineKeyboardButton("❌ NO", callback_data="no"),
    ]])
    await update.message.reply_text(
        "Estimado cliente:\n\n"
        "Por favor, ayúdenos con la cancelación del monto pendiente "
        "correspondiente al servicio prestado.\n\n"
        "*¿Va a realizar el pago?*",
        reply_markup=teclado,
        parse_mode="Markdown"
    )
    return ESPERANDO_RESPUESTA

# ── RAMA SI ──
async def respuesta_si(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data["estado"] = "SI"
    await query.message.reply_text(
        "✅ *¡Gracias por su confirmación!*\n\n"
        "Por favor ingrese la *fecha en que realizará el pago*:\n\n"
        "_Ejemplo: 15/07/2025_",
        parse_mode="Markdown"
    )
    return ESPERANDO_FECHA

async def recibir_fecha(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["fecha"] = update.message.text
    await update.message.reply_text("📋 Por favor ingrese el *nombre de su empresa o negocio*:", parse_mode="Markdown")
    return ESPERANDO_NOMBRE

async def recibir_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["nombre"] = update.message.text
    await update.message.reply_text("🪪 Ingrese su *número de cédula o RUC*:", parse_mode="Markdown")
    return ESPERANDO_CEDULA

async def recibir_cedula(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["cedula"] = update.message.text
    await update.message.reply_text(
        "🧾 Por favor *adjunte la foto o captura del comprobante de pago* y presione enviar 📤",
        parse_mode="Markdown"
    )
    return ESPERANDO_COMPROBANTE

async def recibir_comprobante(update: Update, context: ContextTypes.DEFAULT_TYPE):
    datos = context.user_data
    foto = update.message.photo[-1] if update.message.photo else None
    doc = update.message.document
    usuario = update.effective_user.username or "sin_usuario"

    if not foto and not doc:
        await update.message.reply_text(
            "⚠️ Por favor *adjunte una imagen o archivo* del comprobante de pago.",
            parse_mode="Markdown"
        )
        return ESPERANDO_COMPROBANTE

    nombre_archivo = f"pago_{datos.get('cedula','xxx')}_{datetime.now().strftime('%d%m%Y')}.xlsx"

    # ── A Carolina: mensaje + excel + comprobante ──
    await context.bot.send_message(
        ENCARGADA_ID,
        "📋 *NUEVO REPORTE DE PAGO*\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        f"🏢 Empresa/Nombre: {datos.get('nombre', 'N/A')}\n"
        f"🪪 Cédula/RUC: {datos.get('cedula', 'N/A')}\n"
        f"📅 Fecha de pago: {datos.get('fecha', 'N/A')}\n"
        f"✅ Estado: *VA A PAGAR*\n"
        "━━━━━━━━━━━━━━━━━━━━",
        parse_mode="Markdown"
    )
    await context.bot.send_document(
        ENCARGADA_ID,
        document=generar_excel_pago(datos, usuario),
        filename=nombre_archivo,
        caption="📊 Excel con datos del cliente"
    )
    if foto:
        await context.bot.send_photo(ENCARGADA_ID, foto.file_id, caption="🧾 Comprobante de pago")
    elif doc:
        await context.bot.send_document(ENCARGADA_ID, doc.file_id, caption="🧾 Comprobante de pago")

    # ── Al cliente: su constancia Excel ──
    await context.bot.send_document(
        update.effective_chat.id,
        document=generar_excel_pago(datos, usuario),
        filename=f"constancia_{nombre_archivo}",
        caption="📊 *Su constancia de registro de pago*",
        parse_mode="Markdown"
    )
    await update.message.reply_text(
        "✅ *¡Información recibida correctamente!*\n\n"
        "Le enviamos su constancia en Excel. 📊\n\n"
        "Le recordamos que, en caso de no regularizar su pago dentro del "
        "plazo indicado, el servicio podría ser *suspendido*.\n\n"
        "_Gracias por su atención. 🙏_",
        parse_mode="Markdown"
    )
    context.user_data.clear()
    return ConversationHandler.END

# ── RAMA NO ──
async def respuesta_no(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data["estado"] = "NO"
    await query.message.reply_text(
        "Por favor indique el *motivo* por el cual no realizará el pago:",
        parse_mode="Markdown"
    )
    return ESPERANDO_MOTIVO

async def recibir_motivo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    datos = context.user_data
    datos["motivo"] = update.message.text
    usuario = update.effective_user.username or "sin_usuario"
    nombre_archivo = f"nopago_{update.effective_user.id}_{datetime.now().strftime('%d%m%Y')}.xlsx"

    # ── Solo a Carolina: mensaje + excel ──
    await context.bot.send_message(
        ENCARGADA_ID,
        "🚨 *REPORTE: NO VA A PAGAR*\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 Usuario: @{usuario}\n"
        f"❌ Estado: *NO VA A PAGAR*\n"
        f"📝 Motivo: {datos.get('motivo', 'N/A')}\n"
        "━━━━━━━━━━━━━━━━━━━━",
        parse_mode="Markdown"
    )
    await context.bot.send_document(
        ENCARGADA_ID,
        document=generar_excel_no_pago(datos, usuario),
        filename=nombre_archivo,
        caption="📊 Excel reporte no pago"
    )
    await update.message.reply_text(
        "Le recordamos que, en caso de no regularizar su pago dentro del "
        "plazo establecido, el servicio podría ser *suspendido*.\n\n"
        "Si cambia de opinión, puede comunicarse con nosotros.",
        parse_mode="Markdown"
    )
    context.user_data.clear()
    return ConversationHandler.END

async def cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("Proceso cancelado. Escribe /start para comenzar de nuevo.", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

async def run_bot():
    app = Application.builder().token(TOKEN).build()
    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            ESPERANDO_RESPUESTA: [
                CallbackQueryHandler(respuesta_si, pattern="^si$"),
                CallbackQueryHandler(respuesta_no, pattern="^no$"),
            ],
            ESPERANDO_FECHA:       [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_fecha)],
            ESPERANDO_NOMBRE:      [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_nombre)],
            ESPERANDO_CEDULA:      [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_cedula)],
            ESPERANDO_COMPROBANTE: [MessageHandler(filters.PHOTO | filters.Document.ALL, recibir_comprobante)],
            ESPERANDO_MOTIVO:      [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_motivo)],
        },
        fallbacks=[CommandHandler("cancelar", cancelar)],
        per_message=False,
    )
    app.add_handler(conv)
    logger.info("Bot iniciado ✅")
    await app.initialize()
    await app.start()
    await app.updater.start_polling(drop_pending_updates=True)
    await asyncio.Event().wait()

def main():
    hilo = threading.Thread(target=iniciar_servidor, daemon=True)
    hilo.start()
    def keep_alive():
        import urllib.request, time
        url = os.environ.get("RENDER_EXTERNAL_URL", "")
        while True:
            time.sleep(14 * 60)
            try:
                if url:
                    urllib.request.urlopen(url)
                    logger.info("✅ Ping OK")
            except:
                pass
    threading.Thread(target=keep_alive, daemon=True).start()
    asyncio.run(run_bot())

if __name__ == "__main__":
    main()
